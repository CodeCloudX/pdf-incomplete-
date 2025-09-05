# tools/pdf_to_excel_tool.py
import os
import logging
from flask import current_app
from pdf2docx import Converter
from docx import Document
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from utils.file_utils import validate_file_size
from utils.file_manager import get_session_folder
from utils.file_naming_utils import generate_file_names

# Set up logger
logger = logging.getLogger(__name__)

def pdf_to_excel(file_path, pages=None, table_detection="auto", excel_format="multi"):
    """Convert PDF → Excel preserving text and tables - compatible with generic_tools.py"""
    try:
        # Validate file size
        if not validate_file_size(file_path):
            return {
                'status': 'error',
                'output_files': [],
                'message': f"File {os.path.basename(file_path)} exceeds size limit"
            }

        # Get original filename for display purposes
        original_filename = os.path.basename(file_path)
        
        # Generate secure file names for output
        file_names = generate_file_names(original_filename, toolname='excel', ext='xlsx')
        display_name = file_names['display_name']
        stored_name = file_names['stored_name']

        # Get session-specific processed folder
        processed_folder = get_session_folder('processed')
        out_path = os.path.join(processed_folder, stored_name)
        
        # Get session-specific upload folder for temp files
        upload_folder = get_session_folder('uploads')
        
        # Ensure directories exist
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        os.makedirs(upload_folder, exist_ok=True)

        # Step 1: PDF → Word (temp)
        temp_word_path = os.path.join(upload_folder, f"temp_excel_{stored_name}.docx")
        
        # Convert PDF to Word with optional page selection
        cv = Converter(file_path)
        if pages:
            # Convert specific pages only
            # pdf2docx uses 0-based indexing for start and end
            start_page_0_based = min(pages) - 1
            end_page_0_based = max(pages) - 1
            cv.convert(temp_word_path, start=start_page_0_based, end=end_page_0_based)
        else:
            cv.convert(temp_word_path, start=0, end=None)
        cv.close()

        # Step 2: Word → Excel
        doc = Document(temp_word_path)
        wb = openpyxl.Workbook()
        ws_main = wb.active
        ws_main.title = "PDF Data"

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        row_idx = 1
        table_count = 0
        table_sheet = None

        for block in doc.element.body:
            if block.tag.endswith('p'):  # Paragraphs → main sheet
                para = block.xpath('.//w:t')
                text = ''.join([t.text for t in para if t.text])
                if text.strip():
                    ws_main.cell(row=row_idx, column=1, value=text.strip())
                    ws_main.cell(row=row_idx, column=1).alignment = left_align
                    row_idx += 1

            elif block.tag.endswith('tbl'):  # Tables
                table_count += 1
                if excel_format == "multi":
                    ws_table = wb.create_sheet(title=f"Table_{table_count}")
                else: # single sheet
                    if table_sheet is None:
                        table_sheet = wb.create_sheet(title="All Tables")
                        # Add a header for each table if it's the first table in single mode
                        ws_table.cell(row=ws_table.max_row + 1, column=1, value=f"--- Table {table_count} ---").font = header_font
                        ws_table.cell(row=ws_table.max_row, column=1).alignment = center_align
                    ws_table = table_sheet

                start_row = 1 if excel_format == "multi" else ws_table.max_row + 1

                for r, row in enumerate(block.findall('.//w:tr', namespaces=block.nsmap), start=start_row):
                    for c, cell in enumerate(row.findall('.//w:tc', namespaces=row.nsmap), start=1):
                        texts = [t.text for t in cell.findall('.//w:t', namespaces=cell.nsmap) if t.text]
                        cell_text = " ".join(texts).strip()
                        ws_table.cell(row=r, column=c, value=cell_text)
                        ws_table.cell(row=r, column=c).border = thin_border
                        ws_table.cell(row=r, column=c).alignment = center_align
                        if r == start_row: # Apply header font only to the first row of each table
                            ws_table.cell(row=r, column=c).font = header_font

                # Auto column width
                for col in ws_table.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value and len(str(cell.value)) > max_len:
                                max_len = len(str(cell.value))
                        except:
                            pass
                    ws_table.column_dimensions[col_letter].width = max_len + 2
                
                if excel_format == "single":
                    # Add a blank row after each table in single sheet mode for separation
                    ws_table.cell(row=ws_table.max_row + 1, column=1, value="").border = None

        # Main sheet column width
        for col in ws_main.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws_main.column_dimensions[col_letter].width = max_len + 2

        # Save Excel
        wb.save(out_path)

        # Cleanup temp Word file
        if os.path.exists(temp_word_path):
            os.remove(temp_word_path)

        logger.info(f"Successfully converted {file_path} to Excel: {out_path}")
        
        return {
            'status': 'success',
            'output_files': [{
                'display_name': f"excel_{display_name}.xlsx",
                'stored_name': stored_name,
                'output_path': out_path
            }],
            'message': f'PDF converted to Excel successfully. Table Detection: {table_detection}, Format: {excel_format}'
        }

    except Exception as e:
        logger.error(f"PDF to Excel conversion failed for {file_path}: {str(e)}")
        
        # Cleanup temp files on error
        temp_word_path = os.path.join(get_session_folder('uploads'), f"temp_excel_{os.path.basename(file_path)}.docx")
        if os.path.exists(temp_word_path):
            os.remove(temp_word_path)
            
        return {
            'status': 'error',
            'output_files': [],
            'message': f"Conversion failed: {str(e)}"
        }
